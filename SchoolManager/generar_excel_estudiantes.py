import openpyxl
from datetime import datetime, timedelta
import random

# Crear un nuevo workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Estudiantes"

# Encabezados exactos según la vista
headers = [
    "ESTUDIANTE (EMAIL)",
    "NOMBRE",
    "APELLIDO",
    "DOCUMENTO ID",
    "FECHA NACIMIENTO",
    "GRADO",
    "GRUPO",
    "JORNADA",
    "INCLUSIÓN"
]

# Agregar encabezados
ws.append(headers)

# Estilos para encabezados
from openpyxl.styles import Font, PatternFill, Alignment

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Datos de ejemplo
nombres = ["Juan", "María", "Carlos", "Ana", "Luis", "Laura", "Pedro", "Sofía", "Diego", "Carmen"]
apellidos = ["García", "Rodríguez", "López", "Martínez", "González", "Pérez", "Sánchez", "Ramírez", "Torres", "Flores"]
grados = ["6°", "7°", "8°", "9°", "10°", "11°"]
grupos = ["A", "B", "C", "D", "E"]
jornadas = ["Mañana", "Tarde", "Noche"]
inclusiones = ["si", "no", ""]

# Generar 20 estudiantes de ejemplo
for i in range(20):
    nombre = random.choice(nombres)
    apellido = random.choice(apellidos)
    email = f"{nombre.lower()}.{apellido.lower()}{i+1}@estudiante.com"
    documento = f"EST{i+1:04d}{random.randint(1000, 9999)}"
    
    # Fecha de nacimiento entre 2005 y 2012 (formato DD/MM/YYYY)
    año_nacimiento = random.randint(2005, 2012)
    mes_nacimiento = random.randint(1, 12)
    dia_nacimiento = random.randint(1, 28)
    fecha_nacimiento = f"{dia_nacimiento:02d}/{mes_nacimiento:02d}/{año_nacimiento}"
    
    grado = random.choice(grados)
    grupo = random.choice(grupos)
    jornada = random.choice(jornadas)
    inclusion = random.choice(inclusiones)
    
    ws.append([
        email,
        nombre,
        apellido,
        documento,
        fecha_nacimiento,
        grado,
        grupo,
        jornada,
        inclusion
    ])

# Ajustar ancho de columnas
column_widths = {
    'A': 30,  # ESTUDIANTE (EMAIL)
    'B': 15,  # NOMBRE
    'C': 15,  # APELLIDO
    'D': 15,  # DOCUMENTO ID
    'E': 18,  # FECHA NACIMIENTO
    'F': 10,  # GRADO
    'G': 10,  # GRUPO
    'H': 12,  # JORNADA
    'I': 12   # INCLUSIÓN
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Guardar el archivo
wb.save("asignaciones_estudiantes_grado_grupo.xlsx")
print("Archivo Excel generado exitosamente: asignaciones_estudiantes_grado_grupo.xlsx")
print(f"Total de estudiantes: {20}")

